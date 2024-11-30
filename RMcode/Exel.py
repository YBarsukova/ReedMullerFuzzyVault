import pandas as pd
import os
from openpyxl import load_workbook


def update_excel_with_data(filename, data_lines):
    config = data_lines[0].strip("()").split()
    r, m = int(config[0]), int(config[1])
    column_header = f"Код ({m},{r})"

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        df_initial = pd.DataFrame(columns=["Ошибки\\Коды", column_header])
        df_initial.loc[0] = ["max", 100]
        df_initial.to_excel(filename, index=False)
        workbook = load_workbook(filename)
        sheet = workbook.active

    columns = [cell.value for cell in sheet[1]]
    if column_header not in columns:
        sheet.cell(row=1, column=len(columns) + 1, value=column_header)
        columns.append(column_header)
    col_index = columns.index(column_header) + 1

    for i, line in enumerate(data_lines[1:], start=1):
        _, value = line.split()
        row_label = f"max+{i}"

        row_index = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == row_label:
                row_index = row
                break

        if row_index is None:
            row_index = sheet.max_row + 1
            sheet.cell(row=row_index, column=1, value=row_label)

        percentage_value = round(float(value) * 100)
        sheet.cell(row=row_index, column=col_index, value=percentage_value)

    workbook.save(filename)
